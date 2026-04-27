import customtkinter as ctk
from tkinter import filedialog, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.platypus import Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter
from datetime import datetime
from PIL import ImageGrab, Image
import os
import json
import win32print
import win32api
import threading
import keyboard
import re

# ==============================
# CONFIGURACIÓN
# ==============================

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

archivo_excel = "MODELO.xlsx"
archivo_historial = "clientes_historial.json"
imagen_path = ""
prefijo_factura = "001002000"
facturas_lista = []
combo_impresora = None
combo_copias = None

estilo_grande = ParagraphStyle(
    name="Titulo",
    fontSize=18,
    leading=22,
    alignment=1
)

# ==============================
# HISTORIAL DE CLIENTES
# ==============================

def cargar_historial():
    try:
        with open(archivo_historial, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []

def guardar_en_historial(nombre, cedula):
    historial = cargar_historial()
    # Evitar duplicados por cédula
    historial = [c for c in historial if c.get("cedula") != cedula]
    historial.insert(0, {"nombre": nombre, "cedula": cedula})
    # Mantener solo los últimos 200 clientes
    historial = historial[:200]
    with open(archivo_historial, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)

def buscar_clientes(texto):
    historial = cargar_historial()
    texto = texto.lower().strip()
    if not texto:
        return []
    return [
        c for c in historial
        if texto in c["nombre"].lower() or texto in c.get("cedula", "")
    ][:8]

# ==============================
# AUTOCOMPLETADO
# ==============================

ventana_sugerencias = None

def cerrar_sugerencias():
    global ventana_sugerencias
    if ventana_sugerencias and ventana_sugerencias.winfo_exists():
        ventana_sugerencias.destroy()
        ventana_sugerencias = None

def seleccionar_cliente(cliente):
    entry_nombre.delete(0, "end")
    entry_nombre.insert(0, cliente["nombre"])
    entry_cedula.delete(0, "end")
    entry_cedula.insert(0, cliente.get("cedula", ""))
    cerrar_sugerencias()
    entry_transferencia.focus()

def mostrar_sugerencias(sugerencias):
    global ventana_sugerencias
    cerrar_sugerencias()

    if not sugerencias:
        return

    x = entry_nombre.winfo_rootx()
    y = entry_nombre.winfo_rooty() + entry_nombre.winfo_height()

    ventana_sugerencias = ctk.CTkToplevel(app)
    ventana_sugerencias.wm_overrideredirect(True)
    ventana_sugerencias.geometry(f"400x{min(len(sugerencias)*40, 280)}+{x}+{y}")
    ventana_sugerencias.attributes("-topmost", True)

    for cliente in sugerencias:
        texto = f"{cliente['nombre']}  |  {cliente.get('cedula','')}"
        btn = ctk.CTkButton(
            ventana_sugerencias,
            text=texto,
            anchor="w",
            fg_color="gray20",
            hover_color="gray30",
            height=38,
            command=lambda c=cliente: seleccionar_cliente(c)
        )
        btn.pack(fill="x", padx=2, pady=1)

def on_nombre_keyrelease(event):
    texto = entry_nombre.get()
    if event.keysym == "Escape":
        cerrar_sugerencias()
        return
    sugerencias = buscar_clientes(texto)
    mostrar_sugerencias(sugerencias)

# ==============================
# PEGADO INTELIGENTE
# ==============================

def pegado_inteligente(event=None):
    """
    Intenta parsear el portapapeles como datos del cliente.
    Formatos soportados:
      - Juan Pérez | 1234567890 | 250.00
      - Juan Pérez, 1234567890, 250.00
      - Juan Pérez  1234567890  250.00
    Si no tiene estructura, intenta pegar solo imagen.
    """
    try:
        import tkinter as tk
        texto = app.clipboard_get()
        if texto:
            # Separadores: | , o espacios múltiples o tabs
            partes = re.split(r"[|\t]|,\s*|\s{2,}", texto.strip())
            partes = [p.strip() for p in partes if p.strip()]

            if len(partes) >= 2:
                entry_nombre.delete(0, "end")
                entry_nombre.insert(0, partes[0])

                if len(partes) >= 2:
                    entry_cedula.delete(0, "end")
                    entry_cedula.insert(0, partes[1])

                if len(partes) >= 3:
                    # Intentar poner en transferencia o valor
                    try:
                        float(partes[2])
                        entry_valor.delete(0, "end")
                        entry_valor.insert(0, partes[2])
                    except:
                        entry_transferencia.delete(0, "end")
                        entry_transferencia.insert(0, partes[2])

                mostrar_toast("✅ Datos pegados desde portapapeles")
                return
    except:
        pass

    # Si no hay texto estructurado, intentar imagen
    pegar_imagen()

# ==============================
# INGRESO DE FACTURAS EN LOTE
# ==============================

def agregar_facturas_lote():
    """
    Abre una ventana para pegar múltiples facturas de una vez.
    Formato por línea: 001002000001 | 150.00  o  001002000001 150.00
    """
    ventana_lote = ctk.CTkToplevel(app)
    ventana_lote.title("Agregar facturas en lote")
    ventana_lote.geometry("450x380")
    ventana_lote.attributes("-topmost", True)

    ctk.CTkLabel(
        ventana_lote,
        text="Pega las facturas (una por línea):",
        font=ctk.CTkFont(size=13)
    ).pack(pady=10)

    ctk.CTkLabel(
        ventana_lote,
        text="Formato: 001002000001 | 150.00",
        text_color="gray",
        font=ctk.CTkFont(size=11)
    ).pack()

    txt = ctk.CTkTextbox(ventana_lote, width=400, height=200)
    txt.pack(pady=10, padx=20)

    def procesar_lote():
        contenido = txt.get("1.0", "end").strip()
        lineas = contenido.split("\n")
        agregadas = 0
        errores = []

        for linea in lineas:
            linea = linea.strip()
            if not linea:
                continue

            partes = re.split(r"[|\t,]|\s{2,}|\s", linea)
            partes = [p.strip() for p in partes if p.strip()]

            if len(partes) < 2:
                errores.append(f"Línea inválida: {linea}")
                continue

            factura = partes[0]
            try:
                valor = float(partes[1])
            except:
                errores.append(f"Valor inválido en: {linea}")
                continue

            facturas_lista.append((factura, valor))
            textbox_facturas.insert(
                "end",
                f"Factura: {factura}  |  Valor: ${valor:.2f}\n"
            )
            agregadas += 1

        actualizar_total()
        ventana_lote.destroy()

        msg = f"{agregadas} factura(s) agregadas."
        if errores:
            msg += f"\n\nErrores:\n" + "\n".join(errores)
        messagebox.showinfo("Lote procesado", msg)

    ctk.CTkButton(
        ventana_lote,
        text="Agregar todas",
        fg_color="green",
        command=procesar_lote
    ).pack(pady=10)

# ==============================
# TOAST NOTIFICACIÓN
# ==============================

def mostrar_toast(mensaje, duracion=2000):
    toast = ctk.CTkToplevel(app)
    toast.wm_overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.attributes("-alpha", 0.9)

    w, h = 320, 45
    x = app.winfo_rootx() + (app.winfo_width() // 2) - (w // 2)
    y = app.winfo_rooty() + app.winfo_height() - 80
    toast.geometry(f"{w}x{h}+{x}+{y}")

    ctk.CTkLabel(
        toast,
        text=mensaje,
        font=ctk.CTkFont(size=13),
        fg_color="#1a6b1a",
        corner_radius=8,
        width=w,
        height=h
    ).pack(fill="both", expand=True)

    toast.after(duracion, toast.destroy)

# ==============================
# TOTAL EN TIEMPO REAL
# ==============================

def actualizar_total():
    total = sum(v for _, v in facturas_lista)
    lbl_total.configure(text=f"Total acumulado: ${total:.2f}")

# ==============================
# RESUMEN DEL DÍA
# ==============================

def mostrar_resumen_dia():
    try:
        wb = load_workbook(archivo_excel)
        ws = wb.active
        hoy = datetime.now().strftime("%d/%m/%Y")

        total_dia = 0
        count = 0
        fila = 4
        while ws[f"B{fila}"].value is not None:
            fecha_celda = ws[f"C{fila}"].value
            valor_celda = ws[f"F{fila}"].value
            if str(fecha_celda) == hoy:
                try:
                    total_dia += float(valor_celda)
                    count += 1
                except:
                    pass
            fila += 1

        messagebox.showinfo(
            "Resumen del día",
            f"📅 Fecha: {hoy}\n\n"
            f"📋 Facturas registradas hoy: {count}\n"
            f"💰 Total del día: ${total_dia:.2f}"
        )
    except Exception as e:
        messagebox.showerror("Error", str(e))

# ==============================
# VERIFICAR DUPLICADO
# ==============================

def verificar_duplicado(transferencia):
    try:
        wb = load_workbook(archivo_excel)
        ws = wb.active
        fila = 4
        while ws[f"B{fila}"].value is not None:
            if str(ws[f"B{fila}"].value) == str(transferencia):
                return True
            fila += 1
    except:
        pass
    return False

# ==============================
# CONFIGURACIÓN
# ==============================

def guardar_configuracion():
    config = {
        "impresora": combo_impresora.get(),
        "modo_impresion": combo_copias.get()
    }
    with open("config.json", "w") as f:
        json.dump(config, f)
    messagebox.showinfo("Configuración", "Configuración guardada")

def cargar_configuracion():
    try:
        with open("config.json", "r") as f:
            config = json.load(f)
        combo_impresora.set(config["impresora"])
        combo_copias.set(config["modo_impresion"])
    except:
        pass

def imprimir_pdf(pdf_path):
    try:
        win32api.ShellExecute(0, "print", pdf_path, None, ".", 0)
    except Exception as e:
        print("Error al imprimir:", e)

def obtener_impresoras():
    impresoras = []
    try:
        flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        printers = win32print.EnumPrinters(flags)
        for p in printers:
            impresoras.append(p[2])
    except:
        impresoras = []
    if len(impresoras) == 0:
        impresoras.append("SIN IMPRESORAS")
    return impresoras

def autocompletar_banco(event=None):
    texto = combo_banco.get().upper().strip()
    bancos = {
        "P": "PICHINCHA",
        "G": "GUAYAQUIL",
        "A": "AUSTRO",
        "JA": "JARDIN AZUAYO",
        "J": "JEP",
        "S": "SAN FRANCISCO",
        "PN": "POLICIA NACIONAL"
    }
    if texto in bancos:
        combo_banco.set(bancos[texto])

def seleccionar_imagen():
    global imagen_path
    ruta = filedialog.askopenfilename(
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png")]
    )
    if ruta:
        imagen_path = ruta
        lbl_drop.configure(text="Imagen cargada ✅", text_color="green")

def drop_imagen(event):
    global imagen_path
    archivo = event.data.strip("{}")
    if archivo.lower().endswith((".png", ".jpg", ".jpeg")):
        imagen_path = archivo
        lbl_drop.configure(text="Imagen cargada (drag & drop) ✅", text_color="green")
    else:
        messagebox.showerror("Error", "Solo se permiten imágenes JPG o PNG")

def pegar_imagen(event=None):
    global imagen_path
    img = ImageGrab.grabclipboard()
    if isinstance(img, Image.Image):
        ruta_temp = "temp_comprobante.png"
        img.save(ruta_temp)
        imagen_path = ruta_temp
        lbl_drop.configure(text="Imagen pegada desde portapapeles ✅", text_color="green")
    else:
        messagebox.showerror("Error", "No hay imagen en el portapapeles")

def proteger_prefijo(event):
    texto = entry_factura.get()
    if not texto.startswith(prefijo_factura):
        entry_factura.delete(0, "end")
        entry_factura.insert(0, prefijo_factura)
    if entry_factura.index("insert") < len(prefijo_factura):
        entry_factura.icursor(len(prefijo_factura))

def ordenar_excel(ws):
    datos = []
    fila = 4
    while ws[f"B{fila}"].value is not None:
        factura = ws[f"E{fila}"].value
        try:
            numero_factura = int(str(factura).replace(prefijo_factura, ""))
        except:
            fila += 1
            continue
        datos.append([
            ws[f"B{fila}"].value,
            ws[f"C{fila}"].value,
            ws[f"D{fila}"].value,
            factura,
            ws[f"F{fila}"].value,
            ws[f"G{fila}"].value,
            numero_factura
        ])
        fila += 1

    datos.sort(key=lambda x: x[6])

    fila = 4
    for d in datos:
        ws[f"B{fila}"] = d[0]
        ws[f"C{fila}"] = d[1]
        ws[f"D{fila}"] = d[2]
        ws[f"E{fila}"] = d[3]
        ws[f"F{fila}"] = d[4]
        ws[f"G{fila}"] = d[5]
        fila += 1

def guardar_y_imprimir():
    nombre = entry_nombre.get().strip()
    cedula = entry_cedula.get().strip()
    transferencia = entry_transferencia.get().strip()
    banco = combo_banco.get()
    fecha = datetime.now().strftime("%d/%m/%Y")

    if not nombre or not transferencia or banco == "Seleccione un banco":
        messagebox.showerror("Error", "Complete todos los campos")
        return

    if len(facturas_lista) == 0:
        messagebox.showerror("Error", "Debe agregar al menos una factura")
        return

    # Verificar duplicado
    if verificar_duplicado(transferencia):
        confirmar = messagebox.askyesno(
            "Transferencia duplicada",
            f"⚠️ El número de transferencia '{transferencia}' ya fue registrado.\n\n¿Desea continuar de todas formas?"
        )
        if not confirmar:
            return

    try:
        wb = load_workbook(archivo_excel)
        ws = wb.active

        # Buscar primera fila vacía UNA sola vez
        fila = 4
        while ws[f"B{fila}"].value is not None:
            fila += 1

        # Escribir cada factura incrementando fila
        for factura, valor in facturas_lista:
            ws[f"B{fila}"] = transferencia
            ws[f"C{fila}"] = fecha
            ws[f"D{fila}"] = nombre
            ws[f"E{fila}"] = factura
            ws[f"F{fila}"] = valor
            ws[f"G{fila}"] = banco
            fila += 1

        ordenar_excel(ws)
        wb.save(archivo_excel)

    except Exception as e:
        messagebox.showerror("Error Excel", str(e))
        return

    try:
        total = sum(valor for _, valor in facturas_lista)
        pdf_path = "comprobante.pdf"

        generar_pdf(nombre, cedula, total, facturas_lista, fecha)

        # Guardar cliente en historial
        guardar_en_historial(nombre, cedula)

        # Imprimir directamente sin abrir visor
        imprimir_pdf(pdf_path)

        mostrar_toast("✅ Transferencia registrada e impresa")
        limpiar_campos()

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error:\n{e}")

def limpiar_campos():
    global imagen_path, facturas_lista
    facturas_lista = []

    entry_nombre.delete(0, "end")
    entry_cedula.delete(0, "end")
    entry_transferencia.delete(0, "end")
    entry_valor.delete(0, "end")
    combo_banco.set("Seleccione un banco")
    entry_factura.delete(0, "end")
    entry_factura.insert(0, prefijo_factura)
    textbox_facturas.delete("1.0", "end")
    lbl_drop.configure(text="Arrastra o pega imagen (CTRL+V)", text_color="gray")
    lbl_total.configure(text="Total acumulado: $0.00")
    imagen_path = ""
    cerrar_sugerencias()
    entry_nombre.focus()

def abrir_configuracion():
    ventana = ctk.CTkToplevel(app)
    ventana.title("Configuración")
    ventana.geometry("400x300")

    global combo_impresora, combo_copias

    ctk.CTkLabel(ventana, text="Seleccionar impresora").pack(pady=10)
    combo_impresora = ctk.CTkComboBox(ventana, values=obtener_impresoras(), width=250)
    combo_impresora.pack(pady=10)

    ctk.CTkLabel(ventana, text="Modo de impresión").pack(pady=10)

    if combo_impresora.get() == "SIN IMPRESORAS":
        messagebox.showwarning("Impresora", "No hay impresoras conectadas al sistema.")

    combo_copias = ctk.CTkComboBox(ventana, values=["1 por página", "2 por página"], width=200)
    combo_copias.pack(pady=10)

    ctk.CTkButton(ventana, text="Guardar configuración", command=guardar_configuracion).pack(pady=20)
    cargar_configuracion()

def generar_pdf(nombre, cedula, total, facturas_lista, fecha):
    if not imagen_path:
        raise Exception("Debe agregar una imagen del comprobante")

    ancho, alto = letter
    pdf = SimpleDocTemplate(
        "comprobante.pdf",
        pagesize=letter,
        topMargin=40, bottomMargin=40,
        leftMargin=40, rightMargin=40
    )

    elements = []
    styles = getSampleStyleSheet()
    espacio_imagen = alto - 260

    if imagen_path and os.path.exists(imagen_path):
        img = RLImage(imagen_path)
        max_width = ancho - 80
        max_height = espacio_imagen
        ratio = img.imageWidth / img.imageHeight
        img.drawWidth = max_width
        img.drawHeight = max_width / ratio
        if img.drawHeight > max_height:
            img.drawHeight = max_height
            img.drawWidth = max_height * ratio
        elements.append(img)
        elements.append(Spacer(1, 15))

    elements.append(Paragraph(f"Cliente: {nombre}", estilo_grande))
    elements.append(Paragraph(f"Cédula: {cedula}", estilo_grande))
    elements.append(Paragraph(f"Total: ${total:.2f}", estilo_grande))

    for factura, valor in facturas_lista:
        elements.append(Paragraph(f" {factura}  |  ${valor:.2f}", estilo_grande))

    pdf.build(elements)

def agregar_factura():
    factura = entry_factura.get()
    valor = entry_valor.get()

    if not factura or not valor:
        messagebox.showerror("Error", "Ingrese factura y valor")
        return

    try:
        valor = float(valor)
    except:
        messagebox.showerror("Error", "Valor inválido")
        return

    facturas_lista.append((factura, valor))
    textbox_facturas.insert("end", f"Factura: {factura}  |  Valor: ${valor:.2f}\n")

    entry_factura.delete(0, "end")
    entry_factura.insert(0, prefijo_factura)
    entry_valor.delete(0, "end")

    actualizar_total()
    entry_valor.focus()

def eliminar_ultima_factura():
    if not facturas_lista:
        return
    facturas_lista.pop()
    textbox_facturas.delete("1.0", "end")
    for factura, valor in facturas_lista:
        textbox_facturas.insert("end", f"Factura: {factura}  |  Valor: ${valor:.2f}\n")
    actualizar_total()

# ==============================
# NAVEGACIÓN CON ENTER
# ==============================

def on_enter(event, siguiente=None):
    if siguiente:
        siguiente.focus()

# ==============================
# INTERFAZ
# ==============================

app = TkinterDnD.Tk()
app.title("Sistema de Transferencias")
app.geometry("520x860")

# Atajos globales
app.bind("<Control-v>", pegado_inteligente)
app.bind("<Escape>", lambda e: cerrar_sugerencias())

# ==============================
# BARRA SUPERIOR
# ==============================

barra_superior = ctk.CTkFrame(app)
barra_superior.pack(fill="x", padx=10, pady=5)

btn_config = ctk.CTkButton(
    barra_superior, text="⚙ Impresora",
    command=abrir_configuracion, width=120
)
btn_config.pack(side="left", padx=5)

btn_resumen = ctk.CTkButton(
    barra_superior, text="📊 Resumen día",
    command=mostrar_resumen_dia,
    fg_color="gray30", hover_color="gray40", width=130
)
btn_resumen.pack(side="left", padx=5)

btn_guardar = ctk.CTkButton(
    barra_superior,
    text="💾 Guardar e imprimir",
    fg_color="green", hover_color="#0d5c0d",
    height=35, width=170,
    command=guardar_y_imprimir
)
btn_guardar.pack(side="right", padx=5)

# ==============================
# FORMULARIO (SCROLL)
# ==============================

frame = ctk.CTkScrollableFrame(app, width=500, height=650)
frame.pack(fill="both", expand=True, padx=10, pady=10)

titulo = ctk.CTkLabel(
    frame, text="REGISTRO DE TRANSFERENCIAS",
    font=ctk.CTkFont(size=20, weight="bold")
)
titulo.pack(pady=15)

# Hint pegado inteligente
ctk.CTkLabel(
    frame,
    text="💡 CTRL+V pega datos del cliente o imagen automáticamente",
    text_color="gray", font=ctk.CTkFont(size=11)
).pack(pady=(0, 8))

# Nombre con autocompletado
entry_nombre = ctk.CTkEntry(frame, placeholder_text="Nombre cliente", width=400)
entry_nombre.pack(pady=6)
entry_nombre.bind("<KeyRelease>", on_nombre_keyrelease)
entry_nombre.bind("<Return>", lambda e: on_enter(e, entry_cedula))

entry_cedula = ctk.CTkEntry(frame, placeholder_text="Cédula", width=400)
entry_cedula.pack(pady=6)
entry_cedula.bind("<Return>", lambda e: on_enter(e, entry_transferencia))

entry_transferencia = ctk.CTkEntry(frame, placeholder_text="Número transferencia", width=400)
entry_transferencia.pack(pady=6)
entry_transferencia.bind("<Return>", lambda e: on_enter(e, combo_banco))

lista_bancos = [
    "PICHINCHA", "GUAYAQUIL", "AUSTRO",
    "JARDIN AZUAYO", "JEP", "SAN FRANCISCO", "POLICIA NACIONAL"
]

combo_banco = ctk.CTkComboBox(frame, values=lista_bancos, width=400)
combo_banco.set("Seleccione un banco")
combo_banco.pack(pady=6)
combo_banco.bind("<KeyRelease>", autocompletar_banco)

# ==============================
# SECCIÓN FACTURAS
# ==============================

ctk.CTkLabel(
    frame, text="FACTURAS",
    font=ctk.CTkFont(size=13, weight="bold")
).pack(pady=(12, 2))

frame_factura_row = ctk.CTkFrame(frame, fg_color="transparent")
frame_factura_row.pack(pady=4)

entry_valor = ctk.CTkEntry(frame_factura_row, placeholder_text="Valor", width=130)
entry_valor.pack(side="left", padx=4)

entry_factura = ctk.CTkEntry(frame_factura_row, width=200)
entry_factura.pack(side="left", padx=4)
entry_factura.insert(0, prefijo_factura)

btn_agregar = ctk.CTkButton(
    frame_factura_row, text="+ Agregar",
    width=80, command=agregar_factura
)
btn_agregar.pack(side="left", padx=4)

# Bind Enter en valor para agregar factura
entry_valor.bind("<Return>", lambda e: agregar_factura())
entry_factura.bind("<KeyRelease>", proteger_prefijo)
entry_factura.bind("<Button-1>", proteger_prefijo)

# Botones de lote y eliminar
frame_lote_row = ctk.CTkFrame(frame, fg_color="transparent")
frame_lote_row.pack(pady=2)

btn_lote = ctk.CTkButton(
    frame_lote_row, text="📋 Agregar en lote",
    width=180, fg_color="gray30", hover_color="gray40",
    command=agregar_facturas_lote
)
btn_lote.pack(side="left", padx=6)

btn_eliminar = ctk.CTkButton(
    frame_lote_row, text="✖ Eliminar última",
    width=160, fg_color="#5c1a1a", hover_color="#8b0000",
    command=eliminar_ultima_factura
)
btn_eliminar.pack(side="left", padx=6)

textbox_facturas = ctk.CTkTextbox(frame, width=400, height=90)
textbox_facturas.pack(pady=6)

# Total acumulado
lbl_total = ctk.CTkLabel(
    frame,
    text="Total acumulado: $0.00",
    font=ctk.CTkFont(size=15, weight="bold"),
    text_color="#4CAF50"
)
lbl_total.pack(pady=4)

# ==============================
# IMAGEN
# ==============================

ctk.CTkLabel(
    frame, text="COMPROBANTE",
    font=ctk.CTkFont(size=13, weight="bold")
).pack(pady=(12, 2))

btn_imagen = ctk.CTkButton(frame, text="📁 Seleccionar imagen", command=seleccionar_imagen)
btn_imagen.pack(pady=6)

drop_frame = ctk.CTkFrame(frame, width=400, height=70)
drop_frame.pack(pady=10)
drop_frame.pack_propagate(False)

lbl_drop = ctk.CTkLabel(
    drop_frame,
    text="Arrastra o pega imagen (CTRL+V)",
    text_color="gray"
)
lbl_drop.pack(expand=True)

drop_frame.drop_target_register(DND_FILES)
drop_frame.dnd_bind("<<Drop>>", drop_imagen)

# Focus inicial
entry_nombre.focus()

app.mainloop()
