"""
=============================================================
  API SERVIDOR - PC OFICINA
  Sistema de Transferencias - Firesoft
=============================================================
  Instalar dependencias:
    pip install flask flask-cors openpyxl reportlab pillow

  Ejecutar:
    python api_servidor.py

  La API quedará disponible en:
    http://0.0.0.0:5000
=============================================================
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.platypus import Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter
from datetime import datetime
from PIL import Image
import os
import base64
import json
import win32print
import win32api

app = Flask(__name__)
CORS(app)

# ==============================
# CONFIGURACIÓN
# ==============================

ARCHIVO_EXCEL    = "MODELO.xlsx"
CARPETA_IMAGENES = "imagenes_comprobantes"
CARPETA_PDFS     = "pdfs_generados"
PREFIJO_FACTURA  = "001002000"
ARCHIVO_HISTORIAL = "clientes_historial.json"

os.makedirs(CARPETA_IMAGENES, exist_ok=True)
os.makedirs(CARPETA_PDFS, exist_ok=True)

estilo_grande = ParagraphStyle(
    name="Titulo", fontSize=18, leading=22, alignment=1
)

# ==============================
# UTILIDADES
# ==============================

def ordenar_excel(ws):
    datos = []
    fila = 4
    while ws[f"B{fila}"].value is not None:
        factura = ws[f"E{fila}"].value
        try:
            numero_factura = int(str(factura).replace(PREFIJO_FACTURA, ""))
        except:
            fila += 1
            continue
        datos.append([
            ws[f"B{fila}"].value, ws[f"C{fila}"].value,
            ws[f"D{fila}"].value, factura,
            ws[f"F{fila}"].value, ws[f"G{fila}"].value,
            numero_factura
        ])
        fila += 1
    datos.sort(key=lambda x: x[6])
    fila = 4
    for d in datos:
        ws[f"B{fila}"] = d[0]; ws[f"C{fila}"] = d[1]
        ws[f"D{fila}"] = d[2]; ws[f"E{fila}"] = d[3]
        ws[f"F{fila}"] = d[4]; ws[f"G{fila}"] = d[5]
        fila += 1


def generar_pdf(nombre, cedula, total, facturas_lista, fecha, imagen_path, pdf_path):
    ancho, alto = letter
    pdf = SimpleDocTemplate(pdf_path, pagesize=letter,
        topMargin=40, bottomMargin=40, leftMargin=40, rightMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    espacio_imagen = alto - 260

    if imagen_path and os.path.exists(imagen_path):
        img = RLImage(imagen_path)
        max_width = ancho - 80
        max_height = espacio_imagen
        ratio = img.imageWidth / img.imageHeight
        img.drawWidth = max_width
        img.drawHeight = max_width / ratio
        if img.drawHeight > max_height:
            img.drawHeight = max_height
            img.drawWidth = max_height * ratio
        elements.append(img)
        elements.append(Spacer(1, 15))

    elements.append(Paragraph(f"Fecha: {fecha}", styles["Normal"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Cliente: {nombre}", estilo_grande))
    elements.append(Paragraph(f"Cédula: {cedula}", estilo_grande))
    elements.append(Paragraph(f"Total: ${total:.2f}", estilo_grande))
    for factura, valor in facturas_lista:
        elements.append(Paragraph(f"N° Factura: {factura}  |  Valor: ${valor:.2f}", estilo_grande))
    pdf.build(elements)


def imprimir_pdf(pdf_path):
    try:
        win32api.ShellExecute(0, "print", pdf_path, None, ".", 0)
    except Exception as e:
        print(f"Error al imprimir: {e}")


def guardar_historial(nombre, cedula):
    historial = []
    try:
        with open(ARCHIVO_HISTORIAL, "r", encoding="utf-8") as f:
            historial = json.load(f)
    except:
        pass
    historial = [c for c in historial if c.get("cedula") != cedula]
    historial.insert(0, {"nombre": nombre, "cedula": cedula})
    historial = historial[:200]
    with open(ARCHIVO_HISTORIAL, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)


def verificar_duplicado(transferencia):
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        fila = 4
        while ws[f"B{fila}"].value is not None:
            if str(ws[f"B{fila}"].value) == str(transferencia):
                return True
            fila += 1
    except:
        pass
    return False

# ==============================
# ENDPOINTS
# ==============================

@app.route("/ping", methods=["GET"])
def ping():
    return jsonify({"status": "ok", "mensaje": "Servidor activo ✅"})


@app.route("/clientes", methods=["GET"])
def obtener_clientes():
    try:
        with open(ARCHIVO_HISTORIAL, "r", encoding="utf-8") as f:
            historial = json.load(f)
        return jsonify({"ok": True, "clientes": historial})
    except:
        return jsonify({"ok": True, "clientes": []})


@app.route("/verificar/<transferencia>", methods=["GET"])
def verificar(transferencia):
    duplicado = verificar_duplicado(transferencia)
    return jsonify({"ok": True, "duplicado": duplicado})


@app.route("/resumen", methods=["GET"])
def resumen_dia():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        hoy = datetime.now().strftime("%d/%m/%Y")
        total_dia = 0
        count = 0
        fila = 4
        while ws[f"B{fila}"].value is not None:
            if str(ws[f"C{fila}"].value) == hoy:
                try:
                    total_dia += float(ws[f"F{fila}"].value)
                    count += 1
                except:
                    pass
            fila += 1
        return jsonify({"ok": True, "fecha": hoy, "registros": count, "total": round(total_dia, 2)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/transferencia", methods=["POST"])
def registrar_transferencia():
    """
    Body JSON:
    {
        "nombre": "Juan Pérez",
        "cedula": "1234567890",
        "transferencia": "123456",
        "banco": "PICHINCHA",
        "facturas": [{"factura": "001002000001", "valor": 150.00}],
        "imagen_base64": "data:image/jpeg;base64,..."
    }
    """
    try:
        data = request.get_json()
        nombre        = data.get("nombre", "").strip()
        cedula        = data.get("cedula", "").strip()
        transferencia = data.get("transferencia", "").strip()
        banco         = data.get("banco", "").strip()
        facturas_raw  = data.get("facturas", [])
        imagen_b64    = data.get("imagen_base64", "")

        if not nombre or not transferencia or not banco:
            return jsonify({"ok": False, "error": "Faltan campos obligatorios"}), 400
        if not facturas_raw:
            return jsonify({"ok": False, "error": "Debe incluir al menos una factura"}), 400

        facturas_lista = [(f["factura"], float(f["valor"])) for f in facturas_raw]
        fecha = datetime.now().strftime("%d/%m/%Y")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Guardar imagen
        imagen_path = ""
        if imagen_b64:
            if "," in imagen_b64:
                imagen_b64 = imagen_b64.split(",")[1]
            imagen_bytes = base64.b64decode(imagen_b64)
            imagen_path = os.path.join(CARPETA_IMAGENES, f"comp_{timestamp}.jpg")
            with open(imagen_path, "wb") as img_file:
                img_file.write(imagen_bytes)

        # Escribir en Excel
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        fila = 4
        while ws[f"B{fila}"].value is not None:
            fila += 1
        for factura, valor in facturas_lista:
            ws[f"B{fila}"] = transferencia
            ws[f"C{fila}"] = fecha
            ws[f"D{fila}"] = nombre
            ws[f"E{fila}"] = factura
            ws[f"F{fila}"] = valor
            ws[f"G{fila}"] = banco
            fila += 1
        ordenar_excel(ws)
        wb.save(ARCHIVO_EXCEL)

        # Generar PDF e imprimir
        total = sum(v for _, v in facturas_lista)
        pdf_path = os.path.join(CARPETA_PDFS, f"comp_{timestamp}.pdf")
        generar_pdf(nombre, cedula, total, facturas_lista, fecha, imagen_path, pdf_path)
        imprimir_pdf(pdf_path)
        guardar_historial(nombre, cedula)

        return jsonify({"ok": True, "mensaje": "Registrado correctamente", "total": total})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ==============================
# INICIO
# ==============================

if __name__ == "__main__":
    PUERTO = 8080
    print("=" * 50)
    print("  SERVIDOR TRANSFERENCIAS - Firesoft")
    print("=" * 50)
    print(f"  Excel  : {ARCHIVO_EXCEL}")
    print(f"  Puerto : {PUERTO}")
    print(f"  URL    : http://localhost:{PUERTO}")
    print("=" * 50)
    app.run(host="0.0.0.0", port=PUERTO, debug=False)
